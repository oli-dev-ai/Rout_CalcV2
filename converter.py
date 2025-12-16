import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter

def process_ptv_data(df):
    """
    Przetwarza dane zgodnie z logiką VBA
    """
    # Usuń pierwszy wiersz (stare nagłówki) - VBA zaczyna od wiersza 1 który jest nagłówkiem
    # W pandas pierwszy wiersz to indeks 0
    df_without_headers = df.iloc[1:].reset_index(drop=True)
    
    # Kolumna "Hoja Ruta" (indeks 46, bo Python od 0)
    hoja_ruta_col = 46
    
    # Konwertuj kolumnę na numeric (obsługa stringów i pustych wartości)
    hoja_ruta_series = pd.to_numeric(df_without_headers.iloc[:, hoja_ruta_col], errors='coerce')
    
    # Filtruj wiersze gdzie Hoja Ruta <= 0 lub pusty (NaN po konwersji)
    mask = (hoja_ruta_series.isna()) | (hoja_ruta_series <= 0)
    filtered_df = df_without_headers[mask].copy()
    
    # Kolumny do zachowania (indeksy Python: VBA-1)
    # VBA: 4, 9, 16, 11, 17, 18, 19, 22, 30, 36, 37, 38, 39, 40, 41, 55
    cols_to_keep = [3, 8, 15, 10, 16, 17, 18, 21, 29, 35, 36, 37, 38, 39, 40, 54]
    
    # Docelowe kolumny (indeksy Python: VBA-1)
    # VBA: 48, 27, 4, 5, 6, 8, 7, 49, 50, 21, 20, 22, 23, 24, 25, 26
    target_columns = [47, 26, 3, 4, 5, 7, 6, 48, 49, 20, 19, 21, 22, 23, 24, 25]
    
    # Nagłówki zgodnie z "order set"
    headers = [
        "id", "linked order id", "type", "location id",
        "location name", "location street", "location zip code",
        "location city", "location country", "location group stop time",
        "location stop time", "latitude", "longitude", "loading meters",
        "Capacity 1", "Capacity 2", "Capacity 3", "Capacity 4", "Capacity 5",
        "volume", "weight", "Height", "Width", "Lenght", "pc",
        "support data for stacking factor", "stacking factor",
        "Corrected stacking factor", "Corrected stacking factor 1",
        "Corrected stacking factor 2", "Corrected stacking factor 3",
        "Corrected stacking factor 4", "Corrected stacking factor 5",
        "Corrected stacking factor 6", "service time", "absolute timewindows",
        "time window type", "color", "as is sequence", "tags",
        "forbidden tags", "labels", "penalty cost", "group id",
        "same vehicle group id", "sequence number", "sequence group id", "avis",
        "avis pickup date", "final destination"
    ]
    
    # Utwórz pustą DataFrame z 50 kolumnami
    result_df = pd.DataFrame(index=range(len(filtered_df)), columns=range(50))
    
    # Przepisz dane z oryginalnych kolumn do docelowych
    for src_col, tgt_col in zip(cols_to_keep, target_columns):
        if src_col < len(filtered_df.columns):
            result_df[tgt_col] = filtered_df.iloc[:, src_col].values
    
    # Ustaw nagłówki
    result_df.columns = headers
    
    # Wypełnij stałe wartości
    result_df['type'] = 'PICKUP'
    result_df['location country'] = 'DE'
    result_df['location stop time'] = '1:00'
    result_df['service time'] = '0:01'
    result_df['time window type'] = 'Service Arrival'
    
    return result_df

def add_formulas_to_excel(df, output_path):
    """
    Dodaje formuły Excel do pliku wyjściowego
    """
    # Zapisz DataFrame do Excel
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        
        # Dodaj formuły dla każdego wiersza (pomijając nagłówek)
        for i in range(2, len(df) + 2):  # Excel rows start at 1, +1 for header
            row_num = i
            
            # absolute timewindows (kolumna AJ = 36)
            worksheet[f'AJ{row_num}'].value = f"=IFERROR(VLOOKUP(G{row_num},'https://juliob50600527-my.sharepoint.com/personal/oliwier_opyrchal_gruposese_com/Documents/[opentime.xlsx]sheet'!$B$2:$F$1000,5,FALSE),\"06:00:00-14:00:00\")"
            
            # Latitude (kolumna L = 12)
            worksheet[f'L{row_num}'] = f'=IFERROR(VLOOKUP(E{row_num},\'https://juliob50600527-my.sharepoint.com/personal/oliwier_opyrchal_gruposese_com/Documents/[Points of loading PTV.xlsx]ptvform\'!$A:$D,2,FALSE),"Nie znaleziono")'
            
            # Longitude (kolumna M = 13)
            worksheet[f'M{row_num}'] = f'=IFERROR(VLOOKUP(E{row_num},\'https://juliob50600527-my.sharepoint.com/personal/oliwier_opyrchal_gruposese_com/Documents/[Points of loading PTV.xlsx]ptvform\'!$A:$D,3,FALSE),"Nie znaleziono")'
            
            # Corrected stacking factor (kolumna AB = 28)
            worksheet[f'AB{row_num}'] = f'=AA{row_num}+1'
            
            # Corrected stacking factor 1 (kolumna AC = 29)
            worksheet[f'AC{row_num}'] = f'=IF(ROUNDUP(3/V{row_num},0) > AB{row_num}, AB{row_num}, ROUNDUP(3/V{row_num},0))'

            # Corrected stacking factor 2 (kolumna AD = 30)
            worksheet[f'AD{row_num}'] = f'=IF(ROUNDUP(2.7/V{row_num},0) > AB{row_num}, AB{row_num}, ROUNDUP(2.7/V{row_num},0))'
            
            # Corrected stacking factor 3 (kolumna AE = 31)
            worksheet[f'AE{row_num}'] = f'=IF(ROUNDUP(2.7/V{row_num},0) > AB{row_num}, AB{row_num}, ROUNDUP(2.7/V{row_num},0))'
            
            # Corrected stacking factor 4 (kolumna AF = 32)
            worksheet[f'AF{row_num}'] = f'=IF(ROUNDUP(2.34/V{row_num},0) > AB{row_num}, AB{row_num}, ROUNDUP(2.34/V{row_num},0))'
            
            # Corrected stacking factor 5 (kolumna AG = 33)
            worksheet[f'AG{row_num}'] = f'=IF(ROUNDUP(2.3/V{row_num},0) > AB{row_num}, AB{row_num}, ROUNDUP(2.3/V{row_num},0))'
            
            # Corrected stacking factor 6 (kolumna AH = 34)
            worksheet[f'AH{row_num}'] = f'=IF(ROUNDUP(1.86/V{row_num},0) > AB{row_num}, AB{row_num}, ROUNDUP(1.86/V{row_num},0))'
            
            # Loading meters (kolumna N = 14)
            worksheet[f'N{row_num}'] = f'=IF(OR(AC{row_num}=0, X{row_num}>13.6, W{row_num}>2.48), 999, ROUNDUP(Y{row_num}/AC{row_num},0) * ((W{row_num}*X{row_num})/2.48))'
            
            # Capacity 1 (kolumna O = 15)
            worksheet[f'O{row_num}'] = f'=IF(OR(AD{row_num}=0, X{row_num}>13.6, W{row_num}>2.48), 999, ROUNDUP(Y{row_num}/AD{row_num},0) * ((W{row_num}*X{row_num})/2.48))'
            
            # Capacity 2 (kolumna P = 16)
            worksheet[f'P{row_num}'] = f'=IF(OR(AE{row_num}=0, X{row_num}>7.2, W{row_num}>2.48), 999, ROUNDUP(Y{row_num}/AE{row_num},0) * ((W{row_num}*X{row_num})/2.48))'
            
            # Capacity 3 (kolumna Q = 17)
            worksheet[f'Q{row_num}'] = f'=IF(OR(AF{row_num}=0, X{row_num}>7.2, W{row_num}>2.48), 999, ROUNDUP(Y{row_num}/AF{row_num},0) * ((W{row_num}*X{row_num})/2.48))'
            
            # Capacity 4 (kolumna R = 18)
            worksheet[f'R{row_num}'] = f'=IF(OR(AG{row_num}=0, X{row_num}>4.2, W{row_num}>2.2), 999, ROUNDUP(Y{row_num}/AG{row_num},0) * ((W{row_num}*X{row_num})/2.2))'
            
            # Capacity 5 (kolumna S = 19)
            worksheet[f'S{row_num}'] = f'=IF(OR(AH{row_num}=0, X{row_num}>4, W{row_num}>1.56), 999, ROUNDUP(Y{row_num}/AH{row_num},0) * ((W{row_num}*X{row_num})/1.56))'
            
            # tags (kolumna AN = 40)
            worksheet[f'AN{row_num}'] = f'=IF(E{row_num}="LEM EUROPE GMBH","ML",IF(E{row_num}="IEC GMBH","ML",IF(E{row_num}="ANTARES LIFE CYCLE SOLUTION GMBH","SM",IF(E{row_num}="HEIMSCH DESIGN GMBH","NOBGM",IF(E{row_num}="HENKEL WERK HEIDELBERG","ADR",IF(E{row_num}="PROMENS (ROTOVIA) HOCKENHEIM GMBH","ROT",""))))))'
            
            # labels (kolumna AP = 42)
            worksheet[f'AP{row_num}'] = f'=IF(LEFT(SUBSTITUTE(G{row_num},"DE-",""),1)="5","AREA 1",IF(LEFT(SUBSTITUTE(G{row_num},"DE-",""),1)="6","AREA 2",IF(LEFT(SUBSTITUTE(G{row_num},"DE-",""),1)="7","AREA 3","")))'
            
            # color (kolumna AL = 38)
            worksheet[f'AL{row_num}'] = f'=IF(AP{row_num}="AREA 1","blue",IF(AP{row_num}="AREA 2","green",IF(AP{row_num}="AREA 3","yellow","")))'
            
            # group id (kolumna AS = 44)
            worksheet[f'AS{row_num}'] = f'=AV{row_num}'
        
        # Dodaj obramowania
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows(min_row=1, max_row=len(df)+1, min_col=1, max_col=50):
            for cell in row:
                cell.border = thin_border
        
        # Formatowanie warunkowe dla volume > 99 (kolumna T = 20)
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        white_font = Font(color='FFFFFF')
        
        oversized_rows = []
        for i in range(2, len(df) + 2):
            volume_cell = worksheet[f'T{i}']
            if volume_cell.value and isinstance(volume_cell.value, (int, float)):
                if volume_cell.value > 99:
                    oversized_rows.append(i)
                    # Zaznacz cały wiersz
                    for col in range(1, 51):
                        cell = worksheet.cell(row=i, column=col)
                        cell.fill = red_fill
                        cell.font = white_font
        
        # AutoFit kolumn (przybliżone w openpyxl)
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width


    # Na końcu funkcji, przed return oversized_rows:
    workbook.calculation.calcMode = 'auto'
    workbook.calculation.fullCalcOnLoad = True
    return oversized_rows

# Streamlit App
st.set_page_config(page_title="PTV Data Converter", page_icon="📊", layout="wide")

st.title("🚚 PTV Data Converter")
st.markdown("---")

st.markdown("""
### Instrukcja:
1. Wgraj plik Excel (wsadowy plik przed użyciem VBA)
2. Kliknij **Convert**
3. Pobierz przekonwertowany plik

**Funkcje skryptu:**
- Filtruje dane gdzie `Hoja Ruta <= 0` lub puste
- Reorganizuje kolumny do formatu PTV (50 kolumn)
- Dodaje formuły VLOOKUP, obliczenia i stałe wartości
- Oznacza czerwonym wiersze z `volume > 99`

            
**Ważne**
Po pobraniu pliku i dodaniu etykiety etc. w kolumnie absolute timewindows trzeba enter w pasku formuły.
Pasek formuły to ten biały pasek z tekstem który znajduje się pod paskiem z narzędziami i tam klikamy enter.
Bez tego formuła się nie wczyta, wynika to z błędu excela.
Po wciśnięciu enter, wystarczy przeciągnąć formułę po wszystkich komórkach. """)


st.markdown("---")

# Upload file
uploaded_file = st.file_uploader("📁 Wybierz plik Excel", type=['xlsx', 'xls'])

if uploaded_file is not None:
    try:
        # Wczytaj plik
        df = pd.read_excel(uploaded_file, header=None)
        
        st.success(f"✅ Wczytano plik: **{uploaded_file.name}**")
        st.info(f"📊 Wierszy: {len(df)} | Kolumn: {len(df.columns)}")
        
        # Pokaż preview
        with st.expander("👁️ Podgląd danych wejściowych (pierwsze 5 wierszy)"):
            st.dataframe(df.head())
        
        # Convert button
        if st.button("🔄 Convert", type="primary", use_container_width=True):
            with st.spinner("Przetwarzanie danych..."):
                # Przetwórz dane
                result_df = process_ptv_data(df)
                
                # Zapisz do tymczasowego pliku z formułami
                output_buffer = BytesIO()
                oversized_rows = add_formulas_to_excel(result_df, output_buffer)
                output_buffer.seek(0)
                
                st.success("✅ Konwersja zakończona!")
                
                # Pokaż statystyki
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Wierszy po filtracji", len(result_df))
                with col2:
                    st.metric("Kolumn wyjściowych", 50)
                with col3:
                    st.metric("Ponadnormatywny ładunek", len(oversized_rows))
                
                # Ostrzeżenie o ponadnormatywnym ładunku
                if oversized_rows:
                    st.warning(f"⚠️ **UWAGA!** Znaleziono ponadnormatywny ładunek (volume > 99) w wierszach: {', '.join(map(str, oversized_rows))}")
                
                # Pokaż preview wyniku
                with st.expander("👁️ Podgląd danych wyjściowych (pierwsze 10 wierszy)"):
                    st.dataframe(result_df.head(10))
                
                # Download button
                st.download_button(
                    label="⬇️ Pobierz przekonwertowany plik",
                    data=output_buffer,
                    file_name=f"converted_{uploaded_file.name}",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                
    except Exception as e:
        st.error(f"❌ Błąd podczas przetwarzania: {str(e)}")
        st.exception(e)

else:
    st.info("👆 Wgraj plik Excel aby rozpocząć konwersję")

# Footer
st.markdown("---")
st.markdown("""
<div style='text-align: center; color: gray; font-size: 12px;'>
    PTV Data Converter | Powered by Python & Streamlit
</div>
""", unsafe_allow_html=True)
